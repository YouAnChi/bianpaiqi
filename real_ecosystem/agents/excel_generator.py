import os
import json
import uvicorn
import logging
from pathlib import Path
from starlette.applications import Starlette
from starlette.responses import JSONResponse
from starlette.routing import Route
from langchain_google_genai import ChatGoogleGenerativeAI
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load API Key from environment
from dotenv import load_dotenv
load_dotenv()

if not os.getenv("GOOGLE_API_KEY"):
    raise ValueError("GOOGLE_API_KEY not found in environment variables")

# Configure Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("ExcelGeneratorAgent")

class ExcelGeneratorAgent:
    def __init__(self):
        self.name = "Excel Generator Agent"
        self.system_prompt = """
        你是一位专业的数据表格生成专家。
        你的任务是根据用户需求生成结构化的Excel数据表格。
        你需要：
        1. 理解用户需求，确定表格的主题和数据类型
        2. 生成包含表头和数据行的结构化表格
        3. 数据应该准确、格式清晰、易于理解
        
        请以JSON格式返回表格结构，格式如下：
        {{
            "filename": "文件名",
            "sheets": [
                {{
                    "name": "工作表名称",
                    "headers": ["列1", "列2", "列3"],
                    "data": [
                        ["数据1-1", "数据1-2", "数据1-3"],
                        ["数据2-1", "数据2-2", "数据2-3"]
                    ]
                }}
            ]
        }}
        
        请用中文生成内容，数据要真实合理。
        """
        self.llm = ChatGoogleGenerativeAI(model="gemini-2.0-flash", temperature=0.3)
        self.prompt = ChatPromptTemplate.from_messages([
            ("system", self.system_prompt),
            ("user", "{input}")
        ])
        self.chain = self.prompt | self.llm | StrOutputParser()
        
        # Output directory
        self.output_dir = Path(os.getcwd()) / "output"
        self.output_dir.mkdir(exist_ok=True)

    def create_excel_file(self, excel_structure: dict, filename: str) -> str:
        """根据结构化数据创建Excel文件"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        sheets = excel_structure.get("sheets", [])
        if not sheets:
            # Create a default sheet if none provided
            ws = wb.create_sheet("Sheet1")
            ws.append(["No data provided"])
        else:
            for sheet_data in sheets:
                sheet_name = sheet_data.get("name", "Sheet1")
                ws = wb.create_sheet(sheet_name)
                
                # Add headers
                headers = sheet_data.get("headers", [])
                if headers:
                    ws.append(headers)
                    
                    # Style headers
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Add data rows
                data_rows = sheet_data.get("data", [])
                for row_data in data_rows:
                    ws.append(row_data)
                
                # Auto-adjust column widths
                for col_num in range(1, len(headers) + 1):
                    column_letter = get_column_letter(col_num)
                    max_length = len(str(headers[col_num - 1])) if col_num <= len(headers) else 10
                    
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                        for cell in row:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filepath = self.output_dir / filename
        wb.save(str(filepath))
        logger.info(f"Excel file saved to: {filepath}")
        
        return str(filepath)

    async def handle_request(self, request):
        try:
            body = await request.json()
            logger.info(f"[{self.name}] Received request")
            
            # Extract user message
            user_msg = ""
            if 'params' in body and 'message' in body['params']:
                parts = body['params']['message'].get('parts', [])
                if parts:
                    user_msg = parts[0].get('text', '')
            
            if not user_msg:
                return JSONResponse({"error": "Empty message"}, status_code=400)

            # Try to parse user_msg as JSON (A2A protocol payload)
            task_desc = user_msg
            context = {}
            try:
                payload = json.loads(user_msg)
                if isinstance(payload, dict) and "task_description" in payload:
                    task_desc = payload.get("task_description", "")
                    context = payload.get("context", {})
                    
                    # Reconstruct a better prompt for the LLM
                    formatted_input = f"Task: {task_desc}\n\nContext:\n{json.dumps(context, indent=2, ensure_ascii=False)}"
                    user_msg = formatted_input
                    logger.info(f"[{self.name}] Parsed A2A JSON payload. Task: {task_desc[:50]}...")
            except json.JSONDecodeError:
                pass
            except Exception as e:
                logger.warning(f"[{self.name}] Error parsing payload: {e}")

            logger.info(f"[{self.name}] Processing: {user_msg[:50]}...")
            
            # Call LLM to generate Excel structure
            response_text = await self.chain.ainvoke({"input": user_msg})
            
            logger.info(f"[{self.name}] Response generated ({len(response_text)} chars)")
            
            # Parse JSON structure from response
            try:
                # Try to extract JSON from markdown code blocks
                if "```json" in response_text:
                    json_start = response_text.find("```json") + 7
                    json_end = response_text.find("```", json_start)
                    json_str = response_text[json_start:json_end].strip()
                elif "```" in response_text:
                    json_start = response_text.find("```") + 3
                    json_end = response_text.find("```", json_start)
                    json_str = response_text[json_start:json_end].strip()
                else:
                    json_str = response_text
                
                excel_structure = json.loads(json_str)
                
                # Generate filename
                filename = excel_structure.get("filename", "data.xlsx")
                if not filename.endswith(".xlsx"):
                    filename += ".xlsx"
                
                safe_filename = "".join(
                    [c for c in filename if c.isalnum() or c in ('.', '-', '_')]
                ).strip()
                
                # Create Excel file
                filepath = self.create_excel_file(excel_structure, safe_filename)
                
                sheets_info = excel_structure.get("sheets", [])
                total_rows = sum(len(sheet.get("data", [])) for sheet in sheets_info)
                
                result_text = f"Excel文件已生成！\n\n文件路径: {filepath}\n\n工作表数: {len(sheets_info)}\n总数据行数: {total_rows}"
                
            except json.JSONDecodeError as e:
                logger.warning(f"[{self.name}] Failed to parse JSON, using plain text format: {e}")
                # Fallback: create simple Excel with text
                wb = Workbook()
                ws = wb.active
                ws.title = "数据"
                ws.append(["生成的内容"])
                ws.append([response_text])
                
                filename = "generated_data.xlsx"
                filepath = self.output_dir / filename
                wb.save(str(filepath))
                
                result_text = f"Excel文件已生成（简单格式）！\n\n文件路径: {filepath}\n\n内容:\n{response_text[:200]}..."

            # Construct A2A response
            response_data = {
                "result": {
                    "message": {
                        "role": "model",
                        "parts": [
                            {
                                "text": result_text
                            }
                        ]
                    }
                }
            }
            return JSONResponse(response_data)
            
        except Exception as e:
            logger.error(f"[{self.name}] Error: {e}", exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

agent = ExcelGeneratorAgent()
app = Starlette(debug=True, routes=[
    Route("/", agent.handle_request, methods=["POST"]),
])

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=10009)
